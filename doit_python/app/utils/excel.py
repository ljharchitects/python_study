import os
from pathlib import Path
from typing import List, Any
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font


class Excel:
    def __init__(self, path, dc_rate_percent=None):
        # type: (str, None | int) -> None
        self.path = path
        self.dc_rate_percent = dc_rate_percent

        self.files = self._init_files(path)

    def write(self):
        # type: () -> None
        for file in self.files:
            print(f"READ FILE : {file}")

            path = Path(file)
            output_dir = path.parent.joinpath("after")
            filename = path.name
            output = os.path.join(output_dir, filename)

            workbook = load_workbook(file)
            sheet = workbook["Sheet1"]

            # 디자인
            self._set_design(sheet)

            # 내용 작성
            self._fill_contents(sheet, self.dc_rate_percent)

            # 저장
            workbook.save(output)

    def _init_files(self, path):
        # type: (str) -> List[str]
        files = Path(path).glob("*.xlsx")
        result_list = []
        for file in files:
            result_list.append(file)
        return result_list

    def _fill_contents(self, sheet, dc_rate_percent=None):
        # type: (Worksheet, None | int) -> None
        # 데이터 구하기
        sum_value = self._get_sum(sheet, "C")

        # 내용 작성
        self._fill_content(sheet, 2, 5, "합계금액")
        self._fill_content(sheet, 2, 6, sum_value, "#,###")
        if dc_rate_percent:
            assert dc_rate_percent > 0  # 음수 확인
            dc_rate_float = dc_rate_percent / 100
            self._fill_content(sheet, 3, 5, f"할인 금액({dc_rate_percent}%)")
            self._fill_content(sheet, 3, 6, sum_value * (1 - dc_rate_float), "#,###")

    def _fill_content(self, sheet, row, column, value, number_format=None):
        # type: (Worksheet, int, int, Any, str) -> None
        cell = sheet.cell(row=row, column=column, value=value)
        if number_format:
            cell.number_format = number_format

    def _set_dimension(self, sheet):
        # type: (Worksheet) -> None
        sheet.column_dimensions["A"].width = 11.0
        sheet.column_dimensions["B"].width = 12.0
        sheet.column_dimensions["C"].width = 8.0

    def _get_sum(self, sheet, col_name):
        # type: (Worksheet, str) -> Worksheet
        column = sheet[col_name]
        result = 0
        for cell in column:
            if not cell.value or cell.value == "price":
                continue
            result += int(cell.value)
        return result

    def _set_header(self, sheet):
        # type: (Worksheet) -> None
        row = sheet[1]
        for cell in row:
            cell.font = Font(bold=True)

        sheet.cell(row=2, column=12).font = Font(bold=True)
        sheet.cell(row=2, column=13).font = Font(bold=True)
        sheet.cell(row=3, column=12).font = Font(bold=True, color="00FF0000")
        sheet.cell(row=3, column=13).font = Font(bold=True, color="00FF0000")

    def _set_design(self, sheet):
        # type: (Worksheet) -> None
        self._set_dimension(sheet)
        self._set_header(sheet)
